import os
from openpyxl import Workbook as OPXLWorkbook
from openpyxl import load_workbook as OPXLLoadWorkbook
from typing import Any

from libraries.utils import toList
from libraries.workbook.worksheet import Worksheet

class Workbook:
    """
    A wrapper class for managing OpenPyXL workbooks with enhanced functionality.
    Provides simplified access to workbook operations and worksheet management.
    """

    def __init__(
            self, 
            filename: str | None = None,
            directory: str | None = None,
            worksheet: Any = Worksheet) -> None:
        """
        Initialize a Workbook object.
        
        Args:
            filename: Name of the Excel file
            directory: Directory path where the file is/will be located
            worksheet: Worksheet class to use for sheet creation (defaults to the Worksheet wrapper)
        """
        self._workbook: OPXLWorkbook | None = None
        self._worksheet: Any = worksheet
        self._sheets: dict[str, Worksheet] = dict()
        self._name: str | None = filename
        self._directory: str | None = directory
        self._file: str = os.path.join(directory, filename) if directory and filename else ""
        self._isOpen: bool = False

    # =============================================================================
    # Core Workbook Operations
    # =============================================================================

    def create(self) -> 'Workbook':
        """
        Create a new Excel workbook.
        
        Returns:
            self: For method chaining
        """
        self._workbook = OPXLWorkbook()
        self._isOpen = True
        return self.save()

    def load(self) -> 'Workbook':
        """
        Load an existing Excel workbook from the file path.
        
        Returns:
            self: For method chaining
            
        Raises:
            FileNotFoundError: If the file doesn't exist
        """
        self._workbook = OPXLLoadWorkbook(self._file)
        self._isOpen = True
        return self
    
    def load_data_only(self) -> 'Workbook':
        """
        Load an existing Excel workbook from the file path with data-only mode.
        
        Returns:
            self: For method chaining
        Raises:
            FileNotFoundError: If the file doesn't exist
        """
        self._workbook = OPXLLoadWorkbook(self._file, data_only=True)
        self._isOpen = True
        return self
    
    def save(self) -> 'Workbook':
        """
        Save the workbook to the specified file path.
        
        Returns:
            self: For method chaining
            
        Raises:
            PermissionError: If the file is open or write-protected
        """
        self._workbook.save(self._file)
        return self

    def close(self) -> 'Workbook':
        """
        Close the workbook, releasing system resources.
        
        Returns:
            self: For method chaining
        """
        self._workbook.close()
        self._isOpen = False
        return self
    
    # =============================================================================
    # Worksheet Management
    # =============================================================================

    def hasSheet(self, value: Worksheet | str) -> bool:
        """
        Check if a worksheet exists in the workbook.
        
        Args:
            value: Sheet name (string) or Worksheet object
            
        Returns:
            True if the sheet exists, False otherwise
        """
        if isinstance(value, str):
            return value in self.sheetnames
        return value.name in self.sheetnames

    def openSheet(self, value: Worksheet | str) -> Worksheet | None:
        """
        Open an existing worksheet in the workbook.
        
        Args:
            value: Sheet name (string) or Worksheet object
            
        Returns:
            The requested worksheet or None if not found
        """
        return self._getSheet(value)
    
    def getSheet(self, value: Worksheet | str) -> Worksheet | None:
        """
        Get a worksheet from the workbook.
        
        Args:
            value: Sheet name (string) or Worksheet object
            If a Worksheet object is provided, it will be used directly.
            If a string is provided, the corresponding Worksheet object will be returned.
            If the worksheet does not exist, None will be returned.
            
        Returns:
            The requested worksheet or None if not found
        """
        return self._getSheet(value)
    
    def newSheet(self, value: Worksheet | str) -> Worksheet:
        """
        Create a new worksheet in the workbook.
        
        Args:
            value: Sheet name (string) or Worksheet object
            If a Worksheet object is provided, it will be used directly.
            If a string is provided, a new Worksheet object will be created.
            If the worksheet already exists, it will be returned.
            
        Returns:
            The newly created or existing worksheet
        """
        if not isinstance(value, Worksheet):
            worksheet = Worksheet(value)
        else:
            worksheet = value
        if not self.hasSheet(worksheet):
            self._addSheet(worksheet)
        return worksheet
    
    def insertSheet(self, value: Worksheet | str) -> Worksheet:
        """
        Insert a new worksheet into the workbook.
        
        Args:
            value: Sheet name (string) or Worksheet object
            
        Returns:
            The newly created or existing worksheet
        """
        if isinstance(value, str):
            worksheet: Worksheet = Worksheet(value)
        else:
            worksheet = value
            
        if not self.hasSheet(worksheet):
            self._addSheet(worksheet)
        return worksheet
    
    def deleteSheet(self, value: Worksheet | str) -> 'Workbook':
        """
        Delete a worksheet from the workbook.
        
        Args:
            value: Sheet name (string) or Worksheet object
            
        Returns:
            Self for method chaining
        """
        worksheet: Worksheet | None = self._getSheet(value)
            
        if worksheet is not None:
            self._workbook.remove(self._workbook[worksheet.name])
            if worksheet.name in self._sheets:
                del self._sheets[worksheet.name]
        return self
    
    def orderSheets(self, order: list[str] | None = None) -> 'Workbook':
        """
        Reorder the worksheets in the workbook.
        
        Args:
            order: List of sheet names in the desired order
            
        Returns:
            Self for method chaining
        """
        if order is None:
            order: list[str] | None = self._getSheetsOrder()
            if order is None:
                return self
     
        for item in order:
            item = str(item).strip()
            try:
                change: int = order.index(item) - self.sheetnames.index(item)
                self._workbook.move_sheet(item, offset=change)
            except ValueError as e:
                print(f"Error moving sheet '{item}': {e}")
                return self
        return self
    
    # =============================================================================
    # Properties
    # =============================================================================
    
    @property
    def name(self) -> str | None:
        """Get the name of the workbook file."""
        return self._name
    
    @name.setter
    def name(self, name: str) -> None:
        """
        Set the name of the workbook file.
        
        Args:
            name: New filename for the workbook
        """
        self._name = name
        self._file = os.path.join(self._directory, name) if self._directory else name
    
    @property
    def workbook(self) -> OPXLWorkbook | None:
        """Get the underlying OpenPyXL workbook object."""
        return self._workbook
    
    @workbook.setter
    def workbook(self, workbook: OPXLWorkbook) -> None:
        """
        Set the underlying OpenPyXL workbook object.
        
        Args:
            workbook: An OpenPyXL Workbook object
        """
        self._workbook = workbook

    @property
    def worksheet(self) -> Any:
        """Get the worksheet class used to create new worksheets."""
        return self._worksheet
    
    @worksheet.setter
    def worksheet(self, worksheet: Any) -> None:
        """
        Set the worksheet class used to create new worksheets.
        
        Args:
            worksheet: Worksheet class to use
        """
        self._worksheet = worksheet
    
    @property
    def sheetnames(self) -> list[str]:
        """
        Get a list of all sheet names in the workbook.
        
        Returns:
            List of sheet names as strings
        """
        return self._workbook.sheetnames
    
    @property
    def sheets(self) -> list[Worksheet]:
        """
        Get all worksheet objects in the workbook.
        
        Returns:
            Collection of worksheet objects
        """
        return list(self._sheets.values())
    
    @sheets.setter
    def sheets(self, sheets: Worksheet | list[Worksheet]) -> None:
        """
        Add multiple worksheets to the workbook.
        
        Args:
            sheets: A worksheet or list of worksheets to add
        """
        sheet_list: list[Worksheet] = toList(sheets)
        for sheet in sheet_list:
            if sheet.name not in self.sheetnames:
                self._sheets[sheet.name] = sheet

    @property
    def directory(self) -> str | None:    
        """Get the directory path where the workbook file is located."""
        return self._directory
    
    @directory.setter
    def directory(self, directory: str) -> None:
        """
        Set the directory path where the workbook file is located.
        
        Args:
            directory: Directory path as string
        """
        self._directory = directory
    
    @property
    def file(self) -> str:
        """
        Get the full file path of the workbook.
        
        Returns:
            Full path including directory and filename
        """
        return os.path.join(self._directory, self._name) if self._directory and self._name else ""
    
    @file.setter
    def file(self, filepath: str) -> None:
        """
        Set the full file path of the workbook.
        
        Args:
            filepath: Full path including directory and filename
        """
        self._file = filepath
        self._directory, self._name = os.path.split(filepath)
    
    @property
    def exists(self) -> bool:
        """
        Check if the workbook file exists on disk.
        
        Returns:
            True if the file exists, False otherwise
        """
        return os.path.exists(self.file)
    
    @property
    def isOpen(self) -> bool:
        """
        Check if the workbook is currently open.
        
        Returns:
            True if the workbook is open, False otherwise
        """
        return self._isOpen
    
    @property
    def isEmpty(self) -> bool:
        """
        Check if the workbook is empty (contains only the default sheet).
        
        Returns:
            True if the workbook is empty, False otherwise
        """
        return len(self.sheetnames) == 1 and self.sheetnames[0] == 'Sheet'
    
    # =============================================================================
    # Private Methods
    # =============================================================================
    
    def _addSheet(self, sheet: Worksheet) -> 'Workbook':
        """
        Add a new worksheet to the workbook.
        
        Args:
            sheet: Worksheet object to add
            
        Returns:
            Self for method chaining
        """
        if self.isEmpty:
            self._workbook.remove(self._workbook.active)
       
        self._workbook.create_sheet(title=sheet.name)
        sheet.open(self._workbook[sheet.name])
        self._sheets[sheet.name] = sheet
        return self
    
    def _getSheetsOrder(self) -> list[str] | None:
        """
        Get the custom order for worksheets.
        
        Returns:
            List of sheet names in preferred order, None if default order
        """
        return sorted(self.sheetnames)
    
    def _getSheet(self, value: Worksheet | str) -> Worksheet | None:
        """
        Retrieve a worksheet by name or object.
        
        Args:
            value: Sheet name (string) or Worksheet object
            
        Returns:
            The requested worksheet or None if not found
        """
        if not isinstance(value, str):
            sheet = value
            name = value.name
        else:
            sheet = None
            name = value
      
        if name in self._sheets:
            return self._sheets[name]

        if name in self.sheetnames:
            if sheet is None:
                sheet = Worksheet(name)
            sheet.open(self._workbook[name])
            return sheet

        return None