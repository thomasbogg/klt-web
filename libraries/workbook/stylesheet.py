from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from typing import Literal, Any

from libraries.utils import Object

# Border style types
BorderStyleType = Literal[None, 'thin', 'medium', 'thick', 'dashed', 'dotted', 'double']
# Alignment types
HorizontalAlignType = Literal['left', 'center', 'right', 'justify']
VerticalAlignType = Literal['top', 'center', 'bottom']
# Underline types
UnderlineType = Literal[None, 'single', 'double']

class Stylesheet(Object):
    """
    A wrapper class for managing OpenPyXL style objects.
    Provides a simplified interface for configuring cell styling properties.
    """
        

    def __init__(self, **kwargs) -> None:
        """
        Initialize a Stylesheet with default values.
        
        Default settings:
        - Font: Calibri, 9pt, normal weight, black
        - Alignment: Center horizontal and vertical with text wrapping
        - No borders
        - No fill colors
        """
        values = {
            'fontSize': 9,
            'fontName': 'Calibri',
            'bold': False,
            'italic': False,
            'underline': None,
            'color': '000000',
            'vertical': 'center',
            'horizontal': 'center',
            'wrapText': True,
            'borderRight': None,
            'borderBottom': None,
            'borderLeft': None,
            'borderTop': None,
            'foregroundColour': None,
            'backgroundColour': None,
            'asHeader': False
        }
        values.update(kwargs)
        super().__init__(load=values)

    # =============================================================================
    # Update Method
    # =============================================================================
    
    def update(self, **kwargs) -> None:
        """
        Update the stylesheet with new values.
        
        Args:
            **kwargs: Key-value pairs to update the stylesheet properties
        """
        for key, value in kwargs.items():
            if hasattr(self, key):
                setattr(self, key, value)
            else:
                raise AttributeError(f"Stylesheet has no attribute '{key}'")

    # =============================================================================
    # Style Object Properties
    # =============================================================================

    @property
    def font(self) -> Font:
        """
        Create an OpenPyXL Font object with current font settings.
        
        Returns:
            An OpenPyXL Font object
        """
        return Font(
            name=self.fontName, 
            size=self.fontSize, 
            bold=self.bold, 
            italic=self.italic, 
            underline=self.underline, 
            color=self.color
        )

    @property    
    def alignment(self) -> Alignment:
        """
        Create an OpenPyXL Alignment object with current alignment settings.
        
        Returns:
            An OpenPyXL Alignment object
        """
        return Alignment(
            horizontal=self.horizontal, 
            vertical=self.vertical, 
            wrapText=self.wrapText
        )

    @property
    def border(self) -> Border | None:
        """
        Create an OpenPyXL Border object with current border settings.
        
        Returns:
            An OpenPyXL Border object, or None if no borders are set
        """
        if (
            not self.borderRight and 
            not self.borderBottom and 
            not self.borderLeft and 
            not self.borderTop
        ):
            return None
        
        return Border(
            right=Side(border_style=self.borderRight, style=self.borderRight), 
            bottom=Side(border_style=self.borderBottom, style=self.borderBottom), 
            left=Side(border_style=self.borderLeft, style=self.borderLeft), 
            top=Side(border_style=self.borderTop, style=self.borderTop), 
        )
    
    @border.setter
    def border(self, value: str) -> None:
        """
        Set the border style for the stylesheet.
        Args:
            value: Border style ('thin', 'medium', 'thick', etc.) or None
        """
        self._set('borderRight', value)
        self._set('borderBottom', value)
        self._set('borderLeft', value)
        self._set('borderTop', value)

    @property
    def fill(self) -> PatternFill:
        """
        Create an OpenPyXL PatternFill object with current fill settings.
        
        Returns:
            An OpenPyXL PatternFill object, or a pattern with no fill type if no fill colors are set
        """
        if not self.foregroundColour and not self.backgroundColour: 
            return PatternFill(fill_type=None)
        if not self.foregroundColour:
            return PatternFill(fill_type='solid', fgColor=self.backgroundColour)
        if not self.backgroundColour:
            return PatternFill(fill_type='solid', fgColor=self.foregroundColour)
        return PatternFill(fill_type='solid', fgColor=self.foregroundColour, bgColor=self.backgroundColour)

    # =============================================================================
    # Font Property Getters and Setters
    # =============================================================================
    
    @property
    def fontSize(self) -> int:
        """Get the current font size."""
        return self._get('fontSize')

    @fontSize.setter
    def fontSize(self, value: int) -> None:
        """
        Set the font size.
        
        Args:
            value: Font size in points
        """
        self._set('fontSize', value)

    @property
    def fontName(self) -> str:
        """Get the current font name."""
        return self._get('fontName')
    
    @fontName.setter
    def fontName(self, value: str) -> None:
        """
        Set the font name.
        
        Args:
            value: Font name (e.g., 'Calibri', 'Arial')
        """
        self._set('fontName', value)

    @property
    def bold(self) -> bool:
        """Get the current bold state."""
        return self._get('bold')

    @bold.setter
    def bold(self, value: bool) -> None:
        """
        Set the bold state.
        
        Args:
            value: Boolean indicating if text should be bold
        """
        self._set('bold', value)

    @property
    def italic(self) -> bool:
        """Get the current italic state."""
        return self._get('italic')

    
    @italic.setter
    def italic(self, value: bool) -> None:
        """
        Set the italic state.
        
        Args:
            value: Boolean indicating if text should be italic
        """
        self._set('italic', value)

    @property
    def underline(self) -> UnderlineType:
        """Get the current underline style."""
        return self._get('underline')

    @underline.setter
    def underline(self, value: UnderlineType) -> None:
        """
        Set the underline style.
        
        Args:
            value: Underline style ('single', 'double', None)
        """
        self._set('underline', value)

    @property
    def color(self) -> str:
        """Get the current font color."""
        return self._get('color')
    
    @color.setter
    def color(self, value: str) -> None:
        """
        Set the font color.
        
        Args:
            value: Color in hex format (e.g., '000000' for black)
        """
        self._set('color', value)

    # =============================================================================
    # Alignment Property Getters and Setters
    # =============================================================================

    @property
    def vertical(self) -> VerticalAlignType:
        """Get the current vertical alignment."""
        return self._get('vertical')

    @vertical.setter
    def vertical(self, value: VerticalAlignType) -> None:
        """
        Set the vertical alignment.
        
        Args:
            value: Vertical alignment ('top', 'center', 'bottom')
        """
        self._set('vertical', value)

    @property
    def horizontal(self) -> HorizontalAlignType:
        """Get the current horizontal alignment."""
        return self._get('horizontal')
    
    @horizontal.setter
    def horizontal(self, value: HorizontalAlignType) -> None:
        """
        Set the horizontal alignment.
        
        Args:
            value: Horizontal alignment ('left', 'center', 'right', 'justify')
        """
        self._set('horizontal', value)

    @property
    def wrapText(self) -> bool:
        """Get the current text wrap setting."""
        return self._get('wrapText')

    @wrapText.setter
    def wrapText(self, value: bool) -> None:
        """
        Set whether text should wrap.
        
        Args:
            value: Boolean indicating if text should wrap
        """
        self._set('wrapText', value)

    # =============================================================================
    # Border Property Getters and Setters
    # =============================================================================

    @property
    def borderRight(self) -> BorderStyleType:
        """Get the current right border style."""
        return self._get('borderRight')
    
    @borderRight.setter
    def borderRight(self, value: BorderStyleType) -> None:
        """
        Set the right border style.
        
        Args:
            value: Border style ('thin', 'medium', 'thick', etc.) or None
        """
        self._set('borderRight', value)

    @property
    def borderBottom(self) -> BorderStyleType:
        """Get the current bottom border style."""
        return self._get('borderBottom')
    
    @borderBottom.setter
    def borderBottom(self, value: BorderStyleType) -> None:
        """
        Set the bottom border style.
        
        Args:
            value: Border style ('thin', 'medium', 'thick', etc.) or None
        """
        self._set('borderBottom', value)

    @property
    def borderLeft(self) -> BorderStyleType:
        """Get the current left border style."""
        return self._get('borderLeft')
    
    @borderLeft.setter
    def borderLeft(self, value: BorderStyleType) -> None:
        """
        Set the left border style.
        
        Args:
            value: Border style ('thin', 'medium', 'thick', etc.) or None
        """
        self._set('borderLeft', value)

    @property
    def borderTop(self) -> BorderStyleType:
        """Get the current top border style."""
        return self._get('borderTop')

    @borderTop.setter
    def borderTop(self, value: BorderStyleType) -> None:
        """
        Set the top border style.
        
        Args:
            value: Border style ('thin', 'medium', 'thick', etc.) or None
        """
        self._set('borderTop', value)

    # =============================================================================
    # Fill Color Property Getters and Setters
    # =============================================================================

    @property
    def foregroundColour(self) -> str | None:
        """Get the current foreground color for cell fill."""
        return self._get('foregroundColour')

    @foregroundColour.setter
    def foregroundColour(self, value: str | None) -> None:
        """
        Set the foreground color for cell fill.
        
        Args:
            value: Color in hex format (e.g., 'FFFF00' for yellow)
        """
        self._set('foregroundColour', value)

    @property
    def backgroundColour(self) -> str | None:
        """Get the current background color for cell fill."""
        return self._get('backgroundColour')

    @backgroundColour.setter
    def backgroundColour(self, value: str | None) -> None:
        """
        Set the background color for cell fill.
        
        Args:
            value: Color in hex format (e.g., 'FFFF00' for yellow)
        """
        self._set('backgroundColour', value)